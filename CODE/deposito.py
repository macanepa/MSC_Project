
def Deposito():
    import pandas as pd
    from utilities import Print_Error,get_files,Select_Menu,create_directory



    available_file = []

    search_locations = []
    save_location = ""
    main_config = open("CONFIG/MAIN.config")
    for line in main_config:
        if (line.split(';')[0] == "search_location"):
            search_locations.append(line.split(';')[1].strip())
        elif (line.split(';')[0] == "save_location"):
            save_location=line.split(';')[1].strip()

    for file in get_files(search_locations):
        if(file.upper().endswith('.XLS')):
            available_file.append(file)
    file_name = Select_Menu(available_file,"Select a File",return_type=int)
    file_name = available_file[(file_name)]




    if (file_name.upper().endswith(".XLS")):
        print "Importing XLS File!"
        sheet = "LinnerBooking"
        df = pd.read_excel(io=file_name, sheet_name=sheet)

        df = df[['Booking','Deposito','Weight','Tipo Ctr']]

        df = df.loc[(df['Deposito'] == "MEDLOG SAN ANTONIO") | (df['Deposito'] == "SITRANS SAI ALTO DEPOT")
        | (df['Deposito'] == "SITRANS VALPARAISO DEPOT")
        |(df['Deposito'] == "MEDLOG SANTIAGO")]

        df['Weight'] = df['Weight']/1000 #Transformar a Tons.
        # df = df.loc[(df['Tipo Ctr'] == '20DV') | (df['Tipo Ctr'] == '40DV') | (df['Tipo Ctr'] == '40HC')]
        table = pd.pivot_table(df,values='Weight',aggfunc='count',index='Deposito',columns='Tipo Ctr')
        table = table.reindex(columns=['20DV', '40DV', '40HC'])

        table = table.rename(index={'MEDLOG SAN ANTONIO':'SAI','SITRANS SAI ALTO DEPOT':'SAI',
                                    'SITRANS VALPARAISO DEPOT':'VAP','MEDLOG SANTIAGO':'STGO'})

        table = table.groupby('Deposito').sum()
        # print table.iloc[0]['20DV']



        import openpyxl
        import os

        wb = openpyxl.Workbook()
        sheet = wb.active

        list = []



        print table


        data = []

        for y in range(len(table.index)):
            data.append([])
            for x in range(len(table.columns)):
                data[-1].append(table.iloc[y][x])

        x = 1
        z = 0
        for deposit in data:
            r = 0
            sheet.cell(1,x,str(table.index[z]))
            for value in deposit:
                sheet.cell(2,x,str(table.columns[r]))
                sheet.cell(3,x,float(value))
                x+=1
                r+=1
            x+=1
            z+=1



        wb.save('demo.xlsx')
        wb.close()
        import subprocess




        if (save_location == ""):
            print "Saving Output in Program Location!"
        elif (not os.path.exists(save_location)):
            Print_Error("Save Directory Not Found!")
            create_directory(save_location)


        import platform
        os_platform = platform.system()
        try:
            table.to_excel(save_location+'/file_output.xlsx')
            print "Saved Succesfully"
        except:
            Print_Error('Error Saving File!')

        if(os_platform == 'Linux'):
            try:
                os.system('xdg-open ' + os.getcwd() + '/demo.xlsx')
            except:
                Print_Error("Couldn't open the output file!")

        if(os_platform == 'Windows'):
            try:
                os.system('start ' + os.getcwd() + '/demo.xlsx')
            except:
                Print_Error("Couldn't open the output file!")

    else:
        Print_Error("File not compatible!")
